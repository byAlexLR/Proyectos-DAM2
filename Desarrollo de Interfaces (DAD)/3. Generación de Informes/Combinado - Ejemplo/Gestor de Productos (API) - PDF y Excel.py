#byAlexLR

import tkinter as tk
from tkinter import ttk, messagebox
import requests  # Necesario para comunicar Python con APIs
from fpdf import FPDF  # Biblioteca para generar PDFs
from openpyxl import Workbook  # Biblioteca para generar Excel
from openpyxl.styles import PatternFill, Font, Alignment  # Estilos para Excel
from openpyxl.utils import get_column_letter  # Para ajustar anchos de columnas

class AppTienda:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Productos (API) - PDF y Excel")
        self.root.geometry("900x600")

        # Almacenamiento de datos en memoria
        self.productos_todos = []

        # Interfaz de usuario
        frame_top = tk.Frame(self.root, pady=10)
        frame_top.pack(fill="x")

        # Etiqueta y entrada para filtrar
        tk.Label(frame_top, text="Buscar por Nombre:").pack(side="left", padx=10)

        self.entrada_filtro = tk.Entry(frame_top)
        self.entrada_filtro.pack(side="left", padx=5)
        self.entrada_filtro.bind(
            "<KeyRelease>", self.filtrar_datos
        )  # Filtrar al escribir

        # Botones: Excel y PDF
        btn_pdf = tk.Button(
            frame_top,
            text="Exportar PDF",
            command=self.generar_pdf,
            bg="#FF5722",
            fg="white",
        )
        btn_pdf.pack(side="right", padx=10)

        btn_excel = tk.Button(
            frame_top,
            text="Exportar Excel",
            command=self.generar_excel,
            bg="#217346",
            fg="white",
        )
        btn_excel.pack(side="right", padx=10)

        # Marco principal para la tabla
        frame_tabla = tk.Frame(self.root)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)

        # Definición de columnas
        cols = ("ID", "Título", "Categoría", "Precio", "Stock")
        self.tree = ttk.Treeview(frame_tabla, columns=cols, show="headings")

        # Encabezados de columnas
        self.tree.heading("ID", text="ID")
        self.tree.heading("Título", text="Producto")
        self.tree.heading("Categoría", text="Categoría")
        self.tree.heading("Precio", text="Precio ($)")
        self.tree.heading("Stock", text="Stock")

        # Ancho de columnas
        self.tree.column("ID", width=50, anchor="center")
        self.tree.column("Título", width=300, anchor="center")
        self.tree.column("Categoría", width=150, anchor="center")
        self.tree.column("Precio", width=100, anchor="center")
        self.tree.column("Stock", width=80, anchor="center")

        # Barra de desplazamiento
        scrollbar = ttk.Scrollbar(
            frame_tabla, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Cargar datos desde la API al iniciar
        self.consumir_api()

    # Lógica para consumir API y cargar datos
    def consumir_api(self):
        url = (
            "https://dummyjson.com/products?limit=100"  # API que devuelve 100 productos
        )
        try:
            # Realizamos la petición GET
            respuesta = requests.get(url)
            respuesta.raise_for_status()  # Muestra error si falla
            datos = respuesta.json()

            # Guardamos los productos en nuestra lista
            self.productos_todos = datos["products"]

            # Mostramos los datos en la tabla
            self.actualizar_tabla(self.productos_todos)
        # Manejo de errores de conexión
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error", f"No se pudo conectar a la API:\n{e}")

    # Actualiza la tabla con una lista de datos
    def actualizar_tabla(self, lista_productos):
        # Limpia la tabla
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Inserta los nuevos datos
        for p in lista_productos:
            # Insertar fila en el Treeview
            stock_val = p.get("stock", "")
            self.tree.insert(
                "",
                "end",
                values=(p["id"], p["title"], p["category"], p["price"], stock_val),
            )

    # Lógica para filtrar datos localmente
    def filtrar_datos(self, event=None):
        texto_busqueda = self.entrada_filtro.get().lower()

        if not texto_busqueda:
            # Si el campo está vacío, muestra todos
            self.actualizar_tabla(self.productos_todos)
        else:
            # Filtrar productos por título
            filtrados = [
                p for p in self.productos_todos if texto_busqueda in p["title"].lower()
            ]
            # Actualiza la tabla con los datos filtrados
            self.actualizar_tabla(filtrados)

    # Generar PDF usando FPDF
    def generar_pdf(self):
        # Obtiene todos los datos visibles en la tabla
        items_tabla = self.tree.get_children()

        # Verifica si hay datos para exportar
        if not items_tabla:
            messagebox.showwarning("Aviso", "No hay datos para exportar.")
            return

        # Crea el PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Título del PDF
        pdf.set_font("Arial", "B", 16)
        pdf.cell(200, 10, txt="Informe de Productos Filtrados", ln=True, align="C")
        pdf.ln(10)

        # Encabezado tabla PDF
        pdf.set_font("Arial", "B", 12)
        pdf.set_fill_color(70, 130, 180)  # Azul (Fondo)
        pdf.set_text_color(255, 255, 255)  # Blanco (Texto)
        # Encabezados de tabla en PDF con color de fondo
        pdf.cell(20, 10, "ID", 1, 0, "C", fill=True)
        pdf.cell(100, 10, "Producto", 1, 0, "C", fill=True)
        pdf.cell(40, 10, "Precio", 1, 0, "C", fill=True)
        pdf.cell(30, 10, "Stock", 1, 1, "C", fill=True)

        # Restaurar color negro para las filas
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(0, 0, 0)

        # Filas de datos
        for item in items_tabla:
            valores = self.tree.item(item)["values"]

            # Obtiene y convierte los datos
            id_prod = str(valores[0])
            titulo = str(valores[1])

            # Si el nombre es muy largo, lo trunca
            if len(titulo) > 35:
                titulo = titulo[:32] + "..."

            # Lógica del precio para el formato
            try:
                precio = f"${float(valores[3]):.2f}"
            except:
                precio = str(valores[3])

            # Stock puede venir como int o string
            stock_val = valores[4] if len(valores) > 4 else ""
            try:
                stock_num = int(stock_val)
            except Exception:
                stock_num = None

            # Crea las diferentes celdas
            pdf.cell(20, 10, id_prod, 1, 0, "C")
            pdf.cell(100, 10, titulo, 1, 0, "C")
            pdf.cell(40, 10, precio, 1, 0, "C")

            # Lógica de color para el stock
            if stock_num is None:
                pdf.set_text_color(0, 0, 0)  # Si no es número
            elif stock_num < 50:
                pdf.set_text_color(220, 50, 50)  # Rojo
            else:
                pdf.set_text_color(34, 139, 34)  # Verde

            # Crea la celda de stock
            pdf.cell(30, 10, str(stock_val), 1, 1, "C")

            # Restaura los estilos para la siguiente fila
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "", 10)

        # Guarda la información como un archivo
        nombre_archivo = "informe_productos.pdf"
        try:
            pdf.output(nombre_archivo)
            messagebox.showinfo(
                "Exito", f"PDF generado correctamente:\n{nombre_archivo}"
            )
        # Maneja la posible excepción
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el PDF:\n{e}")

    # Generar Excel usando openpyxl
    def generar_excel(self):
        # Obtiene todos los datos visibles en la tabla
        items_tabla = self.tree.get_children()

        # Verifica si hay datos para exportar
        if not items_tabla:
            messagebox.showwarning("Aviso", "No hay datos para exportar.")
            return

        # Crea el libro y la hoja de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Productos"

        # Definimos los encabezados
        headers = ["ID", "Producto", "Precio", "Stock"]

        # Estilos para el encabezado
        header_fill = PatternFill(
            start_color="4682B4", end_color="4682B4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        # Alineación para encabezado y cuerpo
        center_style = Alignment(horizontal="center", vertical="center")

        # Agrega el encabezado a la hoja
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_style

        # Agrega los datos a la hoja
        for item in items_tabla:
            valores = self.tree.item(item)["values"]
            id_prod = valores[0]
            titulo = valores[1]
            precio = valores[3] if len(valores) > 3 else ""
            stock_val = valores[4] if len(valores) > 4 else ""
            ws.append([id_prod, titulo, precio, stock_val])

        # Colorea las celdas de Stock según la lógica
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
            # Centra todas las celdas
            for cell in row:
                cell.alignment = center_style

            # Lógica específica para la celda de Stock
            stock_cell = row[3]
            try:
                stock_num = int(stock_cell.value)
            except Exception:
                stock_num = None

            # Lógica de color para el stock
            if stock_num is None:
                continue
            elif stock_num < 50:
                stock_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
                stock_cell.font = Font(color="9C0006")
            else:
                stock_cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                stock_cell.font = Font(color="006100")

        # Ajusta el ancho de las columnas
        # 1:ID, 2:Producto, 3:Precio, 4:Stock
        col_widths = {1: 8, 2: 50, 3: 14, 4: 10}
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Guarda los datos como un archivo Excel
        nombre_archivo = "informe_productos.xlsx"
        try:
            wb.save(nombre_archivo)
            messagebox.showinfo(
                "Exito", f"Excel generado correctamente:\n{nombre_archivo}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")

# Ejecución principal
if __name__ == "__main__":
    root = tk.Tk()
    app = AppTienda(root)
    root.mainloop()
