#byAlexLR

import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector  # Necesario para comunicar Python con el servidor MySQL
from fpdf import FPDF  # Biblioteca para generar PDFs
from openpyxl import Workbook  # Biblioteca para generar Excel
from openpyxl.styles import PatternFill, Font, Alignment  # Estilos para Excel
from openpyxl.utils import get_column_letter  # Para ajustar anchos de columnas

# Configuración de la conexión a la base de datos MySQL
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "root",
    "database": "tienda_db",
}

class AppBaseDatos:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Inventario (MySQL) - PDF y Excel")
        self.root.geometry("900x600")

        # Almacenamiento de datos en memoria
        self.datos_memoria = []

        # Interfaz de usuario
        frame_top = tk.Frame(self.root, pady=10)
        frame_top.pack(fill="x")

        # Etiqueta y entrada para filtrar
        tk.Label(frame_top, text="Buscar por Nombre:").pack(side="left", padx=10)

        self.entrada_filtro = tk.Entry(frame_top)
        self.entrada_filtro.pack(side="left", padx=5)
        self.entrada_filtro.bind(
            "<KeyRelease>", self.filtrar_datos
        )  # Filtrar al escribir

        # Botones: Excel y PDF
        btn_pdf = tk.Button(
            frame_top,
            text="Exportar PDF",
            command=self.generar_pdf,
            bg="#FF5722",
            fg="white",
        )
        btn_pdf.pack(side="right", padx=10)

        btn_excel = tk.Button(
            frame_top,
            text="Exportar Excel",
            command=self.generar_excel,
            bg="#217346",
            fg="white",
        )
        btn_excel.pack(side="right", padx=5)

        # Marco principal para la tabla
        frame_tabla = tk.Frame(self.root)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)

        # Definición de columnas
        cols = ("ID", "Nombre", "Categoría", "Precio", "Stock")
        self.tree = ttk.Treeview(frame_tabla, columns=cols, show="headings")

        # Configuración de encabezados y columnas
        for col in cols:
            self.tree.heading(col, text=col)
            if col == "Nombre":
                self.tree.column(col, width=300, anchor="center")
            else:
                self.tree.column(col, width=100, anchor="center")

        # Barra de desplazamiento
        scrollbar = ttk.Scrollbar(
            frame_tabla, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Cargar datos desde MySQL al iniciar
        self.cargar_datos_mysql()

    # Lógica para cargar datos desde MySQL
    def cargar_datos_mysql(self):
        try:
            # Establecer conexión
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)

            # Ejecutar consulta
            cursor.execute("SELECT * FROM productos")
            resultado = cursor.fetchall()

            self.datos_memoria = resultado  # Guardamos los datos en memoria
            self.actualizar_tabla(self.datos_memoria)

            # Cierra la conexión
            cursor.close()
            conn.close()
        # Manejo de errores de conexión
        except mysql.connector.Error as e:
            messagebox.showerror("Error", f"No se pudo conectar a MySQL:\n{e}")

    # Actualiza la tabla con una lista de datos
    def actualizar_tabla(self, lista):
        # Limpia la tabla
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Inserta los nuevos datos
        for fila in lista:
            # Insertar fila en el Treeview
            self.tree.insert(
                "",
                "end",
                values=(
                    fila["id"],
                    fila["nombre"],
                    fila["categoria"],
                    fila["precio"],
                    fila["stock"],
                ),
            )

    # Lógica para filtrar datos localmente
    def filtrar_datos(self, event=None):
        texto = self.entrada_filtro.get().lower()
        if not texto:
            # Si el campo está vacío, muestra todos
            self.actualizar_tabla(self.datos_memoria)
        else:
            # Filtrar productos por nombre
            filtrados = [p for p in self.datos_memoria if texto in p["nombre"].lower()]
            # Actualiza la tabla con los datos filtrados
            self.actualizar_tabla(filtrados)

    # Generar PDF usando FPDF
    def generar_pdf(self):
        # Obtiene todos los datos visibles en la tabla
        items = self.tree.get_children()

        # Verifica si hay datos para exportar
        if not items:
            messagebox.showwarning("Aviso", "No hay datos para exportar.")
            return

        # Crea el PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Título del PDF
        pdf.set_font("Arial", "B", 16)
        pdf.cell(200, 10, txt="Informe de Inventario Filtrado", ln=True, align="C")
        pdf.ln(10)

        # Encabezado tabla PDF
        pdf.set_font("Arial", "B", 12)
        pdf.set_fill_color(70, 130, 180)  # Azul (Fondo)
        pdf.set_text_color(255, 255, 255)  # Blanco (Texto)
        # Encabezados de tabla en PDF con color de fondo
        pdf.cell(15, 10, "ID", border=1, align="C", fill=True)
        pdf.cell(80, 10, "Nombre", border=1, align="C", fill=True)
        pdf.cell(40, 10, "Categoria", border=1, align="C", fill=True)
        pdf.cell(30, 10, "Precio", border=1, align="C", fill=True)
        pdf.cell(25, 10, "Stock", border=1, ln=1, align="C", fill=True)

        # Restaurar color negro para las filas
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(0, 0, 0)

        # Filas de datos
        for item in items:
            valores = self.tree.item(item)["values"]

            # Obtiene y convierte los datos
            id_prod = str(valores[0])
            nombre = str(valores[1])
            categoria = str(valores[2])

            # Si el nombre es muy largo, lo trunca
            if len(nombre) > 35:
                nombre = nombre[:32] + "..."

            # Lógica del precio para el formato
            try:
                precio = f"${float(valores[3]):.2f}"
            except:
                precio = str(valores[3])

            # Comprueba como se devuelve el stock
            stock_val = valores[4] if len(valores) > 4 else ""
            try:
                stock_num = int(stock_val)
            except Exception:
                stock_num = None

            # Crea las diferentes celdas
            pdf.set_text_color(0, 0, 0)
            pdf.cell(15, 10, id_prod, 1, 0, "C")
            pdf.cell(80, 10, nombre, 1, 0, "C")
            pdf.cell(40, 10, categoria, 1, 0, "C")
            pdf.cell(30, 10, precio, 1, 0, "C")

            # Lógica de color para el stock
            if stock_num is None:
                pdf.set_text_color(0, 0, 0)  # Si no es número
            elif stock_num < 50:
                pdf.set_text_color(220, 50, 50)  # Rojo
            else:
                pdf.set_text_color(34, 139, 34)  # Verde

            # Crea la celda de stock
            pdf.cell(25, 10, str(stock_val), 1, 1, "C")

            # Restaura los estilos para la siguiente fila
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "", 10)

        # Guarda la información como un archivo
        nombre_archivo = "informe_inventario.pdf"
        try:
            pdf.output(nombre_archivo)
            messagebox.showinfo(
                "Exito", f"PDF generado correctamente:\n{nombre_archivo}"
            )
        # Maneja la posible excepción
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el PDF:\n{e}")

    # Generar Excel usando Pandas
    def generar_excel(self):
        # Obtiene todos los datos visibles en la tabla
        items = self.tree.get_children()

        # Verifica si hay datos para exportar
        if not items:
            messagebox.showwarning("Aviso", "No hay datos para exportar.")
            return

        # Crea el libro y la hoja de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Inventario"

        # Definimos los encabezados
        headers = ["ID", "Nombre", "Categoría", "Precio", "Stock"]

        # Estilos para el encabezado
        header_fill = PatternFill(
            start_color="4682B4", end_color="4682B4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        # Alineación para encabezado y cuerpo
        center_style = Alignment(horizontal="center", vertical="center")

        # Agrega el encabezado a la hoja
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_style

        # Agrega los datos a la hoja
        for item in items:
            valores = self.tree.item(item)["values"]
            v_id = valores[0]
            v_nombre = valores[1]
            v_cat = valores[2] if len(valores) > 2 else ""
            v_precio = valores[3] if len(valores) > 3 else ""
            v_stock = valores[4] if len(valores) > 4 else ""
            ws.append([v_id, v_nombre, v_cat, v_precio, v_stock])

        # Colorea las celdas de Stock según la lógica
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
            # Centra todas las celdas
            for cell in row:
                cell.alignment = center_style

            # Lógica específica para la celda de Stock
            stock_cell = row[4]
            try:
                stock_num = int(stock_cell.value)
            except Exception:
                stock_num = None

            # Lógica de color para el stock
            if stock_num is not None:
                if stock_num < 50:
                    stock_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    stock_cell.font = Font(color="9C0006")
                else:
                    stock_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    stock_cell.font = Font(color="006100")

        # Ajusta el ancho de las columnas
        # 1:ID, 2:Nombre, 3:Categoría, 4:Precio, 5:Stock
        col_widths = {1: 8, 2: 40, 3: 25, 4: 15, 5: 12}
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Guarda el DataFrame como un archivo Excel
        nombre_archivo = "informe_inventario.xlsx"
        try:
            wb.save(nombre_archivo)
            messagebox.showinfo(
                "Exito", f"Excel generado correctamente:\n{nombre_archivo}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")

# Ejecución principal
if __name__ == "__main__":
    root = tk.Tk()
    app = AppBaseDatos(root)
    root.mainloop()
