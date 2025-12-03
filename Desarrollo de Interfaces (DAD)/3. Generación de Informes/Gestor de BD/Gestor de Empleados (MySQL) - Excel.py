# byAlexLR

import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector  # Necesario para comunicar Python con el servidor MySQL
import pandas as pd  # Biblioteca para generar Excel
from openpyxl.styles import PatternFill, Font  # Para los estilos en Excel

# Configuración de MySQL
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "root",
    "database": "empresa_db",
}


class AppEmpleadosExcel:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Empleados (MySQL) - Excel")
        self.root.geometry("900x600")

        # Almacenamiento de datos en memoria
        self.datos_memoria = []

        # Interfaz de usuario
        frame_top = tk.Frame(self.root, pady=15, padx=10)
        frame_top.pack(fill="x")

        # Etiqueta y entrada para filtrar
        tk.Label(frame_top, text="Buscar Empleado:").pack(side="left", padx=10)
        self.entrada_filtro = tk.Entry(frame_top)
        self.entrada_filtro.pack(side="left", padx=5)
        self.entrada_filtro.bind(
            "<KeyRelease>", self.filtrar_datos
        )  # Filtrar al escribir

        # Botón de exportar a Excel
        btn_excel = tk.Button(
            frame_top,
            text="Exportar a Excel",
            command=self.generar_excel_pandas,
            bg="#FF5722",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        btn_excel.pack(side="right", padx=10)

        # Tabla
        frame_tabla = tk.Frame(self.root)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)

        # Columnas de la tabla
        cols = ("ID", "Nombre", "Puesto", "Salario", "Fecha Contratación")
        self.tree = ttk.Treeview(frame_tabla, columns=cols, show="headings")

        # Configuración de encabezados y columnas
        self.tree.heading("ID", text="ID")
        self.tree.heading("Nombre", text="Nombre")
        self.tree.heading("Puesto", text="Puesto")
        self.tree.heading("Salario", text="Salario ($)")
        self.tree.heading("Fecha Contratación", text="Fecha Contratación")

        self.tree.column("ID", width=50, anchor="center")
        self.tree.column("Nombre", width=200)
        self.tree.column("Puesto", width=200, anchor="center")
        self.tree.column("Salario", width=100, anchor="center")
        self.tree.column("Fecha Contratación", width=150, anchor="center")

        # Para desplazarse por la tabla
        scrollbar = ttk.Scrollbar(
            frame_tabla, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Carga los datos iniciales
        self.cargar_datos_mysql()

    # Lógica para cargar datos desde MySQL
    def cargar_datos_mysql(self):
        try:
            # Establece la conexión
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)

            # Consulta para obtener todos los empleados
            cursor.execute("SELECT * FROM empleados")

            # Guardamos como lista de tuplas
            self.datos_memoria = cursor.fetchall()
            self.actualizar_tabla(self.datos_memoria)

            # Cierra la conexión
            cursor.close()
            conn.close()
        # Captura errores de conexión o consulta
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", f"No se pudo conectar a la BD: {e}")

    # Actualiza la tabla con una lista
    def actualizar_tabla(self, lista):
        # Limpia la tabla
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Inserta los nuevos datos
        for p in lista:
            fecha_str = str(p["fecha_contratacion"])

            self.tree.insert(
                "",
                "end",
                values=(
                    p["id"],
                    p["nombre"],
                    p["puesto"],
                    f"${p['salario']}",
                    fecha_str,
                ),
            )

    # Lógica para filtrar datos localmente
    def filtrar_datos(self, event=None):
        texto = self.entrada_filtro.get().lower()
        if not texto:
            # Si el campo está vacío, muestra todos
            self.actualizar_tabla(self.datos_memoria)
        else:
           # Filtra por nombre o puesto
            filtrados = [
                p
                for p in self.datos_memoria
                if texto in str(p["nombre"]).lower()
                or texto in str(p["puesto"]).lower()
            ]
            # Actualiza la tabla con los datos filtrados
            self.actualizar_tabla(filtrados)

    # Exporta a Excel usando Pandas
    def generar_excel_pandas(self):
        # Verifica si hay datos en la tabla
        if not self.datos_memoria:
            messagebox.showwarning("Aviso", "No hay datos para exportar")
            return

        try:
            # Crea DataFrame principal
            df = pd.DataFrame(self.datos_memoria)

            # Verifica que las columnas esperadas estén presentes
            columnas_requeridas = ["nombre", "puesto", "salario", "fecha_contratacion"]
            if not all(col in df.columns for col in columnas_requeridas):
                messagebox.showerror(
                    "Error",
                    f"Las columnas de la base de datos no coinciden.\nSe esperaban: {columnas_requeridas}\nSe encontraron: {list(df.columns)}",
                )
                return

            # Limpia y renombra columnas para el Excel
            df_export = df[columnas_requeridas].copy()
            df_export.columns = ["Nombre", "Puesto", "Salario", "Fecha Contratación"]

            # Asegura que el salario sea numérico
            df_export["Salario"] = pd.to_numeric(df_export["Salario"], errors="coerce")

            # Crea DataFrame resumen con el salario promedio por puesto de trabajo
            df_resumen = (
                df_export.groupby("Puesto")["Salario"].mean().round(2).reset_index()
            )
            df_resumen.columns = ["Puesto", "Salario Promedio"]

            nombre_archivo = "informe_empleados.xlsx"

            # Exporta usando Pandas con múltiples hojas y estilos
            with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:

                # La hoja principal con el listado completo
                df_export.to_excel(writer, index=False, sheet_name="Listado Empleados")

                # La hoja resumen con el salario promedio por puesto de trabajo
                df_resumen.to_excel(writer, index=False, sheet_name="Resumen Salarios")

                # Estilos para los encabezados
                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_font = Font(color="FFFFFF", bold=True)

                # Aplica los estilos a todas las hojas
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    # Aplica los estilos a la fila de encabezados
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                    # Ajusta el ancho de las columnas
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 25

            messagebox.showinfo(
                "Exito", f"Excel generado correctamente:\n{nombre_archivo}"
            )
        # Captura errores al guardar el Excel
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")


# Ejecuta la aplicación
if __name__ == "__main__":
    root = tk.Tk()
    app = AppEmpleadosExcel(root)
    root.mainloop()

